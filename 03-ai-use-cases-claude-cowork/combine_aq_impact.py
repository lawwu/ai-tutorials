"""
combine_aq_impact.py
────────────────────────────────────────────────────────────────────────────────
Merges Audience Quality (AQ) and Impact data from an Excel workbook into:
  1. A formatted Excel workbook  (AQ_Impact_Combined.xlsx)
  2. An interactive HTML dashboard (AQ_Impact_Dashboard.html)

USAGE
-----
  python combine_aq_impact.py                          # uses default input file
  python combine_aq_impact.py "My Campaign Data.xlsx"  # pass a custom file path

INPUT FILE REQUIREMENTS
-----------------------
The input Excel workbook must contain two sheets named exactly:
  • "AQ"     — Audience Quality data
  • "Impact" — Impact data

Both sheets must share these dimension columns (used as join keys):
  Time Period, Time Stamp, Contracted Data Break, Publisher,
  Publisher Type, Ad Type, Placement Detail, Targeting

AQ sheet must contain:
  AQ Segments, Consumer Reach, Target Reach, AQI, Frequency, Confidence

Impact sheet must contain:
  SOB, SOB Type, Incremental Conversions, Relative Lift,
  Frequency to Impact, Stat Sig
"""

import sys
import json
import warnings
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ── Configuration ──────────────────────────────────────────────────────────────
DEFAULT_INPUT = "Shannon Claude Test.xlsx"
OUTPUT_XLSX   = "AQ_Impact_Combined.xlsx"
OUTPUT_HTML   = "AQ_Impact_Dashboard.html"

JOIN_KEYS = [
    'Time Period', 'Time Stamp', 'Contracted Data Break',
    'Publisher', 'Publisher Type', 'Ad Type', 'Placement Detail', 'Targeting'
]

AQ_DIM      = ['AQ Segments']
AQ_METRICS  = ['Consumer Reach', 'Target Reach', 'AQI', 'AQ Frequency', 'Confidence']
IMP_DIM     = ['SOB', 'SOB Type']
IMP_METRICS = ['Incremental Conversions', 'Relative Lift', 'Frequency to Impact', 'Stat Sig']

# ── Colours ────────────────────────────────────────────────────────────────────
C_NAVY   = 'FF2E4057'
C_BLUE   = 'FF1B6CA8'
C_TEAL   = 'FF0D7377'
C_EVEN   = 'FFF0F4F8'
C_WHITE  = 'FFFFFFFF'
C_DARK   = 'FF1A1A2E'
C_MUTED  = 'FFFFFFFF'


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Load & merge
# ══════════════════════════════════════════════════════════════════════════════
def load_and_merge(path):
    aq  = pd.read_excel(path, sheet_name='AQ')
    imp = pd.read_excel(path, sheet_name='Impact')
    aq.columns  = [c.strip() for c in aq.columns]
    imp.columns = [c.strip() for c in imp.columns]
    aq = aq.rename(columns={'Frequency': 'AQ Frequency'})

    merged = pd.merge(aq, imp, on=JOIN_KEYS, how='outer')
    cols   = JOIN_KEYS + AQ_DIM + AQ_METRICS + IMP_DIM + IMP_METRICS
    merged = merged[cols].sort_values(
        by=['Contracted Data Break', 'Publisher', 'Ad Type', 'AQ Segments', 'SOB', 'SOB Type'],
        na_position='last'
    ).reset_index(drop=True)
    return aq, imp, merged


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Build Excel
# ══════════════════════════════════════════════════════════════════════════════
def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _font(color='FFFFFFFF', sz=10, bold=True):
    return Font(name='Arial', bold=bold, size=sz, color=color)

def _border():
    s = Side(style='thin', color='FFD3D3D3')
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _left():
    return Alignment(horizontal='left', vertical='center')

FRIENDLY = {
    'Time Period': 'Time\nPeriod', 'Time Stamp': 'Time\nStamp',
    'Contracted Data Break': 'Contracted Data\nBreak',
    'Publisher': 'Publisher', 'Publisher Type': 'Publisher\nType',
    'Ad Type': 'Ad Type', 'Placement Detail': 'Placement\nDetail',
    'Targeting': 'Targeting', 'AQ Segments': 'AQ Segment',
    'Consumer Reach': 'Consumer\nReach', 'Target Reach': 'Target\nReach',
    'AQI': 'AQI', 'AQ Frequency': 'Frequency', 'Confidence': 'Confidence',
    'SOB': 'SOB', 'SOB Type': 'SOB Type',
    'Incremental Conversions': 'Incremental\nConversions',
    'Relative Lift': 'Relative\nLift',
    'Frequency to Impact': 'Freq to\nImpact', 'Stat Sig': 'Stat Sig',
}

COL_WIDTHS = {
    'Time Period': 8, 'Time Stamp': 9, 'Contracted Data Break': 22,
    'Publisher': 18, 'Publisher Type': 12, 'Ad Type': 12,
    'Placement Detail': 14, 'Targeting': 11, 'AQ Segments': 18,
    'Consumer Reach': 12, 'Target Reach': 11, 'AQI': 8,
    'AQ Frequency': 10, 'Confidence': 11, 'SOB': 12, 'SOB Type': 10,
    'Incremental Conversions': 14, 'Relative Lift': 10,
    'Frequency to Impact': 11, 'Stat Sig': 9,
}


def _add_sheet_data(wb, title, df, hdr_color):
    ws = wb.create_sheet(title)
    ws.row_dimensions[1].height = 28
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = _fill(hdr_color); c.font = _font(sz=10); c.alignment = _center(); c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 18)
    for ri, row in df.iterrows():
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri+2, column=ci, value=None if pd.isna(val) else val)
            c.font = _font(color=C_DARK, bold=False, sz=10)
            c.alignment = _left(); c.border = _border()
            c.fill = _fill(C_EVEN) if ri % 2 == 0 else _fill(C_WHITE)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(df.columns))}{len(df)+1}'
    return ws


def build_excel(aq_orig, imp_orig, merged, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Combined View'

    cols   = list(merged.columns)
    n_dim  = len(JOIN_KEYS)
    n_aqd  = len(AQ_DIM)
    n_aqm  = len(AQ_METRICS)
    n_impd = len(IMP_DIM)

    aq_start  = n_dim + 1
    aq_end    = n_dim + n_aqd + n_aqm
    imp_start = aq_end + 1
    imp_end   = len(cols)

    # Row 1 – section banners
    ws.row_dimensions[1].height = 18
    for c in range(1, n_dim + 1):
        ws.cell(row=1, column=c).fill = _fill(C_NAVY)
        ws.cell(row=1, column=c).font = _font(sz=9)
        ws.cell(row=1, column=c).alignment = _center()

    ws.merge_cells(start_row=1, start_column=aq_start, end_row=1, end_column=aq_end)
    c = ws.cell(row=1, column=aq_start, value='◆  AUDIENCE QUALITY')
    c.fill = _fill(C_BLUE); c.font = _font(sz=10); c.alignment = _center()
    for ci in range(aq_start+1, aq_end+1):
        ws.cell(row=1, column=ci).fill = _fill(C_BLUE)

    ws.merge_cells(start_row=1, start_column=imp_start, end_row=1, end_column=imp_end)
    c = ws.cell(row=1, column=imp_start, value='◆  IMPACT')
    c.fill = _fill(C_TEAL); c.font = _font(sz=10); c.alignment = _center()
    for ci in range(imp_start+1, imp_end+1):
        ws.cell(row=1, column=ci).fill = _fill(C_TEAL)

    # Row 2 – column headers
    ws.row_dimensions[2].height = 32
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=2, column=ci, value=FRIENDLY.get(col, col))
        cell.alignment = _center(); cell.border = _border()
        if ci <= n_dim:
            cell.fill = _fill('FF3D5472'); cell.font = _font(sz=9)
        elif ci <= aq_end:
            cell.fill = _fill('FF2682C5'); cell.font = _font(sz=9)
        else:
            cell.fill = _fill('FF11969B'); cell.font = _font(sz=9)
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 12)

    # Data rows
    for ri, row in merged.iterrows():
        excel_row = ri + 3
        ws.row_dimensions[excel_row].height = 16
        row_fill = _fill(C_EVEN) if ri % 2 == 0 else _fill(C_WHITE)
        for ci, col in enumerate(cols, 1):
            val = row[col]
            val = None if pd.isna(val) else val
            cell = ws.cell(row=excel_row, column=ci, value=val)
            cell.font  = _font(color=C_DARK, bold=False, sz=10)
            cell.border = _border(); cell.fill = row_fill
            if col in ('Consumer Reach', 'Target Reach', 'Incremental Conversions'):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col in ('AQI', 'Relative Lift'):
                cell.number_format = '0.000'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col in ('AQ Frequency', 'Frequency to Impact'):
                cell.number_format = '0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = _left()

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(cols))}{len(merged)+2}'

    # Original sheets
    aq_orig2  = aq_orig.rename(columns={'AQ Frequency': 'Frequency'}, errors='ignore')
    _add_sheet_data(wb, 'AQ Data',     aq_orig2,  C_BLUE)
    _add_sheet_data(wb, 'Impact Data', imp_orig,  C_TEAL)

    wb.save(out_path)
    print(f"  ✓  Excel saved → {out_path}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Build HTML dashboard
# ══════════════════════════════════════════════════════════════════════════════
def build_html(merged, out_path):
    df = merged.where(pd.notna(merged), None)
    data_json   = json.dumps(df.to_dict(orient='records'))

    def uniq(col):
        return sorted([str(v) for v in df[col].dropna().unique() if v])

    filter_json = json.dumps({
        'contracted': uniq('Contracted Data Break'),
        'publisher':  uniq('Publisher'),
        'adtype':     uniq('Ad Type'),
        'aqseg':      uniq('AQ Segments'),
        'sob':        uniq('SOB'),
        'sobtype':    uniq('SOB Type'),
    })

    # Inline the full HTML template (same as original build)
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Campaign Health Dashboard — AQ &amp; Impact</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
  :root {{
    --navy:#2E4057;--blue:#1B6CA8;--teal:#0D7377;--light:#F0F4F8;
    --white:#FFFFFF;--border:#D3D3D3;--text:#1A1A2E;--muted:#6B7280;
    --aq-bg:#D0E8F7;--aq-hdr:#1B6CA8;--imp-bg:#D0F0EE;--imp-hdr:#0D7377;
  }}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:Arial,sans-serif;background:#f5f7fa;color:var(--text);font-size:13px}}
  header{{background:var(--navy);color:white;padding:16px 24px;display:flex;align-items:center;gap:16px}}
  header h1{{font-size:18px;font-weight:700}}
  header span{{font-size:12px;opacity:.7}}
  .filters{{background:white;border-bottom:1px solid var(--border);padding:12px 24px;display:flex;flex-wrap:wrap;gap:12px;align-items:flex-end}}
  .filter-group{{display:flex;flex-direction:column;gap:4px}}
  .filter-group label{{font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.5px}}
  .filter-group select{{border:1px solid var(--border);border-radius:4px;padding:5px 8px;font-size:12px;color:var(--text);background:white;min-width:140px}}
  .btn-reset{{padding:6px 14px;background:var(--navy);color:white;border:none;border-radius:4px;cursor:pointer;font-size:12px;font-weight:600;align-self:flex-end}}
  .kpi-bar{{display:flex;gap:12px;padding:14px 24px;flex-wrap:wrap}}
  .kpi{{background:white;border:1px solid var(--border);border-radius:6px;padding:10px 16px;min-width:130px;flex:1}}
  .kpi .label{{font-size:10px;color:var(--muted);text-transform:uppercase;font-weight:700;margin-bottom:4px}}
  .kpi .value{{font-size:20px;font-weight:700;color:var(--text)}}
  .kpi .sub{{font-size:11px;color:var(--muted)}}
  .kpi.aq{{border-top:3px solid var(--blue)}}.kpi.imp{{border-top:3px solid var(--teal)}}
  .charts{{display:flex;gap:16px;padding:0 24px 16px;flex-wrap:wrap}}
  .chart-card{{background:white;border:1px solid var(--border);border-radius:6px;padding:16px;flex:1;min-width:280px;max-width:420px}}
  .chart-card h3{{font-size:12px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:12px}}
  .table-wrap{{padding:0 24px 24px;overflow-x:auto}}
  .result-count{{font-size:12px;color:var(--muted);margin-bottom:8px}}
  table{{width:100%;border-collapse:collapse;font-size:12px}}
  thead tr.section-row th{{padding:5px 8px;font-size:11px;font-weight:700;color:white;text-align:center;border:1px solid rgba(255,255,255,.3)}}
  thead tr.section-row .dim{{background:var(--navy)}}.aq{{background:var(--aq-hdr)}}.imp{{background:var(--imp-hdr)}}
  thead tr.col-header th{{padding:6px 8px;font-size:11px;font-weight:700;color:white;text-align:center;border:1px solid rgba(255,255,255,.3);white-space:pre-line;line-height:1.3}}
  thead tr.col-header .dim{{background:#3d5472}}
  thead tr.col-header .aq{{background:#2682c5}}
  thead tr.col-header .imp{{background:#11969b}}
  tbody tr{{background:white}}tbody tr:nth-child(even){{background:var(--light)}}tbody tr:hover{{background:#e8f0fb}}
  tbody td{{padding:6px 8px;border:1px solid #e8e8e8;white-space:nowrap}}
  tbody td.aq-col{{background-color:rgba(208,232,247,.15)}}tbody td.imp-col{{background-color:rgba(208,240,238,.15)}}
  tbody tr:nth-child(even) td.aq-col{{background-color:rgba(208,232,247,.30)}}tbody tr:nth-child(even) td.imp-col{{background-color:rgba(208,240,238,.30)}}
  tbody td.num{{text-align:right}}
  .badge{{display:inline-block;padding:2px 7px;border-radius:10px;font-size:10px;font-weight:700}}
  .badge-high{{background:#d1fae5;color:#065f46}}.badge-medium{{background:#fef3c7;color:#92400e}}.badge-low{{background:#fee2e2;color:#991b1b}}
  .no-rows{{text-align:center;padding:40px;color:var(--muted);font-style:italic}}
</style>
</head>
<body>
<header><div><h1>Campaign Health Dashboard</h1><span>Audience Quality &amp; Impact — Holistic View</span></div></header>
<div class="filters">
  <div class="filter-group"><label>Contracted Data Break</label><select id="f-contracted"><option value="">All</option></select></div>
  <div class="filter-group"><label>Publisher</label><select id="f-publisher"><option value="">All</option></select></div>
  <div class="filter-group"><label>Ad Type</label><select id="f-adtype"><option value="">All</option></select></div>
  <div class="filter-group"><label>AQ Segment</label><select id="f-aqseg"><option value="">All</option></select></div>
  <div class="filter-group"><label>SOB</label><select id="f-sob"><option value="">All</option></select></div>
  <div class="filter-group"><label>SOB Type</label><select id="f-sobtype"><option value="">All</option></select></div>
  <button class="btn-reset" onclick="resetFilters()">Reset</button>
</div>
<div class="kpi-bar">
  <div class="kpi aq"><div class="label">Avg AQI</div><div class="value" id="kpi-aqi">—</div><div class="sub">Audience Quality Index</div></div>
  <div class="kpi aq"><div class="label">Avg Target Reach</div><div class="value" id="kpi-reach">—</div><div class="sub">Target audience reach</div></div>
  <div class="kpi aq"><div class="label">Top Confidence</div><div class="value" id="kpi-conf">—</div><div class="sub">AQ confidence level</div></div>
  <div class="kpi imp"><div class="label">Total Incr. Conversions</div><div class="value" id="kpi-conv">—</div><div class="sub">Incremental conversions</div></div>
  <div class="kpi imp"><div class="label">Avg Relative Lift</div><div class="value" id="kpi-lift">—</div><div class="sub">Relative lift</div></div>
  <div class="kpi imp"><div class="label">Avg Freq to Impact</div><div class="value" id="kpi-freq">—</div><div class="sub">Frequency to impact</div></div>
</div>
<div class="charts">
  <div class="chart-card"><h3>AQI by AQ Segment</h3><canvas id="chart-aqi" height="180"></canvas></div>
  <div class="chart-card"><h3>Incremental Conversions by SOB</h3><canvas id="chart-conv" height="180"></canvas></div>
  <div class="chart-card"><h3>Target Reach by Publisher</h3><canvas id="chart-reach" height="180"></canvas></div>
</div>
<div class="table-wrap">
  <div class="result-count" id="row-count"></div>
  <table id="main-table">
    <thead>
      <tr class="section-row">
        <th class="dim" colspan="8">DIMENSIONS</th>
        <th class="aq" colspan="6">AUDIENCE QUALITY</th>
        <th class="imp" colspan="6">IMPACT</th>
      </tr>
      <tr class="col-header">
        <th class="dim">Time\\nPeriod</th><th class="dim">Time\\nStamp</th><th class="dim">Contracted\\nData Break</th>
        <th class="dim">Publisher</th><th class="dim">Publisher\\nType</th><th class="dim">Ad Type</th>
        <th class="dim">Placement\\nDetail</th><th class="dim">Targeting</th>
        <th class="aq">AQ Segment</th><th class="aq">Consumer\\nReach</th><th class="aq">Target\\nReach</th>
        <th class="aq">AQI</th><th class="aq">Frequency</th><th class="aq">Confidence</th>
        <th class="imp">SOB</th><th class="imp">SOB Type</th><th class="imp">Incremental\\nConversions</th>
        <th class="imp">Relative\\nLift</th><th class="imp">Freq to\\nImpact</th><th class="imp">Stat Sig</th>
      </tr>
    </thead>
    <tbody id="table-body"></tbody>
  </table>
</div>
<script>
const ALL_DATA={data_json};
const FILTERS={filter_json};
function populateSelect(id,vals){{const s=document.getElementById(id);vals.forEach(v=>{{const o=document.createElement('option');o.value=v;o.textContent=v;s.appendChild(o)}})}}
populateSelect('f-contracted',FILTERS.contracted);populateSelect('f-publisher',FILTERS.publisher);
populateSelect('f-adtype',FILTERS.adtype);populateSelect('f-aqseg',FILTERS.aqseg);
populateSelect('f-sob',FILTERS.sob);populateSelect('f-sobtype',FILTERS.sobtype);
['f-contracted','f-publisher','f-adtype','f-aqseg','f-sob','f-sobtype'].forEach(id=>document.getElementById(id).addEventListener('change',refresh));
function getF(){{return{{contracted:document.getElementById('f-contracted').value,publisher:document.getElementById('f-publisher').value,adtype:document.getElementById('f-adtype').value,aqseg:document.getElementById('f-aqseg').value,sob:document.getElementById('f-sob').value,sobtype:document.getElementById('f-sobtype').value}}}}
function filterData(){{const f=getF();return ALL_DATA.filter(r=>(!f.contracted||r['Contracted Data Break']===f.contracted)&&(!f.publisher||r['Publisher']===f.publisher)&&(!f.adtype||r['Ad Type']===f.adtype)&&(!f.aqseg||r['AQ Segments']===f.aqseg)&&(!f.sob||r['SOB']===f.sob)&&(!f.sobtype||r['SOB Type']===f.sobtype))}}
function badge(v){{if(!v)return'';const l=v.toLowerCase();if(l==='high')return`<span class="badge badge-high">${{v}}</span>`;if(l==='medium')return`<span class="badge badge-medium">${{v}}</span>`;if(l==='low')return`<span class="badge badge-low">${{v}}</span>`;return v}}
function fmt(v,t){{if(v===null||v===undefined||v==='')return'<span style="color:#ccc">—</span>';if(t==='comma')return Number(v).toLocaleString();if(t==='dec3')return Number(v).toFixed(3);if(t==='int')return Math.round(Number(v));return v}}
function avg(arr,k){{const v=arr.map(r=>r[k]).filter(v=>v!==null&&v!==undefined&&v!=='');return v.length?(v.reduce((a,b)=>a+Number(b),0)/v.length):null}}
let cAqi=null,cConv=null,cReach=null;
function updateKPIs(rows){{
  const aqi=avg(rows,'AQI'),reach=avg(rows,'Target Reach'),conv=rows.reduce((s,r)=>s+(r['Incremental Conversions']!==null&&r['Incremental Conversions']!==undefined?Number(r['Incremental Conversions']):0),0),lift=avg(rows,'Relative Lift'),freq=avg(rows,'Frequency to Impact');
  const cc={{}};rows.forEach(r=>{{if(r['Confidence'])cc[r['Confidence']]=(cc[r['Confidence']]||0)+1}});const tc=Object.entries(cc).sort((a,b)=>b[1]-a[1])[0];
  document.getElementById('kpi-aqi').textContent=aqi!==null?aqi.toFixed(2):'—';
  document.getElementById('kpi-reach').textContent=reach!==null?Math.round(reach).toLocaleString():'—';
  document.getElementById('kpi-conf').textContent=tc?tc[0]:'—';
  document.getElementById('kpi-conv').textContent=conv>0?Math.round(conv).toLocaleString():'—';
  document.getElementById('kpi-lift').textContent=lift!==null?lift.toFixed(3):'—';
  document.getElementById('kpi-freq').textContent=freq!==null?freq.toFixed(1):'—';
}}
function updateCharts(rows){{
  const aqiBySeq={{}};rows.forEach(r=>{{if(r['AQ Segments']&&r['AQI']!==null){{if(!aqiBySeq[r['AQ Segments']])aqiBySeq[r['AQ Segments']]=[];aqiBySeq[r['AQ Segments']].push(Number(r['AQI']))}}}});
  const al=Object.keys(aqiBySeq),av2=al.map(k=>(aqiBySeq[k].reduce((a,b)=>a+b,0)/aqiBySeq[k].length).toFixed(2));
  if(cAqi)cAqi.destroy();cAqi=new Chart(document.getElementById('chart-aqi'),{{type:'bar',data:{{labels:al,datasets:[{{data:av2,backgroundColor:'#1B6CA8',borderRadius:4}}]}},options:{{plugins:{{legend:{{display:false}}}},scales:{{y:{{beginAtZero:true}}}},responsive:true}}}});
  const cvBySob={{}};rows.forEach(r=>{{if(r['SOB']&&r['Incremental Conversions']!==null)cvBySob[r['SOB']]=(cvBySob[r['SOB']]||0)+Number(r['Incremental Conversions'])}});
  const cl=Object.keys(cvBySob),cv=cl.map(k=>cvBySob[k]);
  if(cConv)cConv.destroy();cConv=new Chart(document.getElementById('chart-conv'),{{type:'bar',data:{{labels:cl,datasets:[{{data:cv,backgroundColor:'#0D7377',borderRadius:4}}]}},options:{{plugins:{{legend:{{display:false}}}},scales:{{y:{{beginAtZero:true}}}},responsive:true}}}});
  const rpub={{}};rows.forEach(r=>{{if(r['Publisher']&&r['Target Reach']!==null){{if(!rpub[r['Publisher']])rpub[r['Publisher']]=[];rpub[r['Publisher']].push(Number(r['Target Reach']))}}}});
  const rl=Object.keys(rpub),rv=rl.map(k=>Math.round(rpub[k].reduce((a,b)=>a+b,0)/rpub[k].length));
  if(cReach)cReach.destroy();cReach=new Chart(document.getElementById('chart-reach'),{{type:'bar',data:{{labels:rl,datasets:[{{data:rv,backgroundColor:'#2E4057',borderRadius:4}}]}},options:{{plugins:{{legend:{{display:false}}}},scales:{{y:{{beginAtZero:true}}}},responsive:true}}}});
}}
function renderTable(rows){{
  const tb=document.getElementById('table-body');
  document.getElementById('row-count').textContent=`Showing ${{rows.length}} row${{rows.length!==1?'s':''}}`;
  if(!rows.length){{tb.innerHTML='<tr><td colspan="20" class="no-rows">No rows match the selected filters.</td></tr>';return}}
  tb.innerHTML=rows.map(r=>`<tr>
    <td>${{r['Time Period']||''}}</td><td>${{r['Time Stamp']||''}}</td><td><strong>${{r['Contracted Data Break']||''}}</strong></td>
    <td>${{r['Publisher']||''}}</td><td>${{r['Publisher Type']||''}}</td><td>${{r['Ad Type']||''}}</td>
    <td>${{r['Placement Detail']||''}}</td><td>${{r['Targeting']||''}}</td>
    <td class="aq-col"><strong>${{r['AQ Segments']||'<span style=\\"color:#ccc\\">—</span>'}}</strong></td>
    <td class="aq-col num">${{fmt(r['Consumer Reach'],'comma')}}</td><td class="aq-col num">${{fmt(r['Target Reach'],'comma')}}</td>
    <td class="aq-col num">${{fmt(r['AQI'],'dec3')}}</td><td class="aq-col num">${{fmt(r['AQ Frequency'],'int')}}</td>
    <td class="aq-col">${{badge(r['Confidence'])}}</td>
    <td class="imp-col">${{r['SOB']||'<span style=\\"color:#ccc\\">—</span>'}}</td><td class="imp-col">${{r['SOB Type']||'<span style=\\"color:#ccc\\">—</span>'}}</td>
    <td class="imp-col num">${{fmt(r['Incremental Conversions'],'comma')}}</td><td class="imp-col num">${{fmt(r['Relative Lift'],'dec3')}}</td>
    <td class="imp-col num">${{fmt(r['Frequency to Impact'],'int')}}</td><td class="imp-col">${{badge(r['Stat Sig'])}}</td>
  </tr>`).join('');
}}
function refresh(){{const rows=filterData();updateKPIs(rows);updateCharts(rows);renderTable(rows)}}
function resetFilters(){{['f-contracted','f-publisher','f-adtype','f-aqseg','f-sob','f-sobtype'].forEach(id=>document.getElementById(id).value='');refresh()}}
refresh();
</script>
</body>
</html>"""

    with open(out_path, 'w') as f:
        f.write(html)
    print(f"  ✓  HTML saved    → {out_path}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    input_file = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT
    print(f"\nProcessing: {input_file}")
    print("─" * 50)

    aq_orig, imp_orig, merged = load_and_merge(input_file)
    print(f"  ✓  Merged {len(merged)} rows from {len(aq_orig)} AQ + {len(imp_orig)} Impact rows")

    build_excel(aq_orig, imp_orig, merged, OUTPUT_XLSX)
    build_html(merged, OUTPUT_HTML)

    print("─" * 50)
    print("Done! Both outputs are in the same folder as this script.\n")
